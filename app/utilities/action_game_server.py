
import os
import time
import threading
import tkinter as tk
import xlwings as xw
import xlwings as xw
import json
import openpyxl
from openpyxl.utils import get_column_letter
from loguru import logger
from bokeh.io import output_file, show
from bokeh.layouts import column, row
from bokeh.plotting import figure
from bokeh.models import CheckboxGroup, ColumnDataSource, CustomJS, Label

#Fixed

from app.utilities.Inventory_DB import InventoryDatabase
from app.settings.configs import UTF_8

class ActionGameServers(object):
    def __init__(self, **kwargs):
        self.Inv = InventoryDatabase()
        #config_db
        self.database = json.load(open("app/config/database_config.json"))
        self.Inv.config = json.load(open("app/config/game_config.json"))
        self.Inv.database_init(self.database)
        self.Inv.create_table()
        self.demand = []

        if isinstance(self.Inv, Exception):
            logger.error(self.Inv)
            raise Exception(self.Inv)
        # SET DURATION
        try:
            self.duration = self.Inv.config['game_period']
        except Exception as e:
            self.duration = 0

        self.utf_8 = UTF_8
        self.screen = kwargs.get('screen', None)
        self.connection_game = kwargs.get('connection_game', None)
        self.CONNECTION = kwargs.get('CONNECTION', {})
        self.DISCONNECTED_PLAYERS = kwargs.get('DISCONNECTED_PLAYERS', [])
        self.loginthread = kwargs.get('loginthread', None)
        self.position = kwargs.get('position', 0)

        self.path_save = os.path.abspath("./save")
        self.check_directory_is_create(self.path_save)

        GAME_STATE = "START"

        self.policy = 0
        self.init_position = 0
        self.receivethread = {}
        self.outputthread = {}
        self.EOQ = {}
        self.check_players_name = []
        self.instance = None


    # def get_sheet_by_name(self, name):
    #     if name == "A":
    #         return self.Inv[self.sheet_a], self.Inv_data_only[self.sheet_a]
    #     elif name == "B":
    #         return self.Inv[self.sheet_b], self.Inv_data_only[self.sheet_b]
    #     elif name == "RESULT":
    #         return self.Inv[self.sheet_result], self.Inv_data_only[self.sheet_result]
    #     elif name == "CONFIG":
    #         return self.Inv[self.sheet_config], self.Inv_data_only[self.sheet_config]


    def server_receive_order(self, client_id):
        player = self.CONNECTION[client_id]["name"]
        logger.debug(f"server_receive_order | waiting_recv_ordering : {player}")
        logger.debug(f"server_receive_order | self.CONNECTION: {self.CONNECTION}")
        try:
            receive = self.CONNECTION[client_id]["client_socket"].recv(16).decode(self.utf_8)
            logger.debug(f"server_receive_order | recv_ordering : {player}")
            if receive == "ordered":
                self.screen.receive_order(player)
                self.CONNECTION[client_id]["client_socket"].send("closed".encode(self.utf_8))
                logger.debug(f"server_receive_order | ordered send_command_closed : {player}")
            elif receive == "not ordered":
                self.CONNECTION[client_id]["client_socket"].send("closed".encode(self.utf_8))
                logger.debug(f"server_receive_order | not ordered send_command_closed: {player}")
        except ConnectionResetError as e:
            self.DISCONNECTED_PLAYERS.append(client_id)
            name = self.CONNECTION[client_id]["name"]
            # clear connection and loginthread
            self.receivethread = {}
            del self.CONNECTION[client_id]
            del self.loginthread[client_id]
            logger.error(f"server_receive_order | ConnectionResetError: {player} | self.CONNECTION: {self.CONNECTION}")
            logger.error(f"server_receive_order | DISCONNECTED_PLAYERS: {self.DISCONNECTED_PLAYERS}")

            # Call waiting_reconnect function and start server_wait_order again
            time.sleep(2)
            response_reconnect = self.connection_game.waiting_reconnect(client_id,name)
            if isinstance(response_reconnect, Exception):
                logger.error(f"server_receive_order | response_reconnect: {response_reconnect}")
                raise Exception(response_reconnect)
            self.screen.mainscreen()
            self.server_wait_order()
        
        logger.debug("server_receive_order | thread closed")

    def server_wait_order(self):
        for client_id in self.CONNECTION:
            self.receivethread[client_id] = threading.Thread(
                target = self.server_receive_order,
                args = (client_id,),
            )
        
        logger.debug("server_wait_order | waiting_recv_ready")
        [self.receivethread[client_id].start() for client_id in self.receivethread]

    def server_next_turn(self, skip=0):
        order = {}
        self.screen.widgets["nextButton"].config(state=tk.DISABLED)
        self.screen.widgets["skipButton"].config(state=tk.DISABLED)
        self.screen.clear_order()
        self.Inv.demand_order_create()
    

        for client_id in self.CONNECTION:
            player_name = self.CONNECTION[client_id]["name"]
            self.CONNECTION[client_id]["client_socket"].send("next".encode(self.utf_8))
            logger.debug(f"server_next_turn | send_command_next_{player_name}")

            # Inv_data for get value or excel formula | Inv_data_only for get value without formula
            
            # Inv_data, Inv_data_only = self.get_sheet_by_name(player_name)
            # Inv_data = {
            #         "CONFIG": sheet_config,
            #         "A": sheet_a,
            #         "B": sheet_b,
            #         "RESULT": sheet_result
            #     } 
            self.receivethread[client_id].join()
            if self.policy:
                logger.debug(f"server_next_turn | waiting_recv_ready_{player_name}")
                ready = self.CONNECTION[client_id]["client_socket"].recv(16).decode(self.utf_8)
                logger.debug(f"server_next_turn | recv_ready_{player_name}")
                self.CONNECTION[client_id]["client_socket"].send(str(skip).encode(self.utf_8))
                logger.debug(f"server_next_turn | send_skip_{player_name}")

                # Inv_data[f"L{self.position + 4}"].value = Inv_data[(2+self.position, 12)].value

            else:
                logger.debug(f"server_next_turn | waiting_recv_order_{player_name}")
                order[player_name] = self.CONNECTION[client_id]["client_socket"].recv(16).decode(self.utf_8)
                logger.debug(f"server_next_turn | recv_order_{player_name}")
                
                # Inv_data["L"+str(4 +self.position)].value = order[player_name]
                self.Inv.create_new_order(order[player_name],player_name)
                # """



        current_row_A = self.Inv.get_current_row("A")
        current_row_B = self.Inv.get_current_row("B")
        current_demand = self.Inv.get_current_demand()

        self.screen.edit_score(
            self.position,
            "Ordering",
            A=[current_row_A[self.Inv.col_dict_inv["order"]]],
            B=[current_row_B[self.Inv.col_dict_inv["order"]]],
        ) if self.position - self.init_position else None
        

         
        self.Inv.update_cashflow_detail("A")
        self.Inv.update_cashflow_detail("B")
        self.screen.set_total(
            profit=[self.Inv.INV["A"]["cash_flow"]["Sales Profit"],self.Inv.INV["B"]["cash_flow"]["Sales Profit"]],
            # # fix=[-int(self.Inv[team]["X12"].value) for team in self.CONNECTION],
            var=[-self.Inv.INV["A"]["cash_flow"]["Variable Cost"],-self.Inv.INV["B"]["cash_flow"]["Variable Cost"]],
        )
        
        self.position = self.position + 1

        # self.Inv_data["F15"].value = self.position
        self.demand.append(current_demand)

        '''
        Create data for sending to interface 
        column number come from Inventory_DB self.col_dict_inv
        '''
        column = [0,1,5,8,2,7,11,12,13,14,15]
        A = []
        B = []
        
        for col in column:
            A.append(current_row_A[col])
            B.append(current_row_B[col])
        print("False12")
        # Inv = xw.Book(self.excel_db.path_db).sheets
        self.screen.add_score(self.position,skip=skip,A=A,B=B,Main=[current_row_A[0],self.demand[-1]])
        print("False13")
        self.Inv.update_cashflow_detail("A")
        self.Inv.update_cashflow_detail("B")

        self.screen.update()
        print("Finish turn Process")

        for client_id in self.CONNECTION:
            player_name = self.CONNECTION[client_id]["name"]
            # Inv_data for get value or excel formula | Inv_data_only for get value without formula
            if player_name == "A":
                data = {
                    'period': self.position,
                    'demand': current_row_A[self.Inv.col_dict_inv["demand"]],
                    'onhand': current_row_A[self.Inv.col_dict_inv["ending_inv"]],
                    'onorder': current_row_A[self.Inv.col_dict_inv["on_order"]],
                    'arrive': current_row_A[self.Inv.col_dict_inv["arrive"]],
                    'receive': current_row_A[self.Inv.col_dict_inv["recieve"]],
                    'sold': current_row_A[self.Inv.col_dict_inv["sold"]],
                    'Shortage': current_row_A[self.Inv.col_dict_inv["shortage"]],
                    'profits': self.Inv.INV[player_name]["cash_flow"]["Sales Profit"],
                    'varcost': -self.Inv.INV[player_name]["cash_flow"]["Variable Cost"],
                }
            else:
                data = {
                    'period': self.position,
                    'demand': current_row_B[self.Inv.col_dict_inv["demand"]],
                    'onhand': current_row_B[self.Inv.col_dict_inv["ending_inv"]],
                    'onorder': current_row_B[self.Inv.col_dict_inv["on_order"]],
                    'arrive': current_row_B[self.Inv.col_dict_inv["arrive"]],
                    'receive': current_row_B[self.Inv.col_dict_inv["recieve"]],
                    'sold': current_row_B[self.Inv.col_dict_inv["sold"]],
                    'Shortage': current_row_B[self.Inv.col_dict_inv["shortage"]],
                    'profits': self.Inv.INV[player_name]["cash_flow"]["Sales Profit"],
                    'varcost': -self.Inv.INV[player_name]["cash_flow"]["Variable Cost"],
                }
            output_send = str(data).encode('UTF-8')
            self.CONNECTION[client_id]["client_socket"].send(output_send)
            logger.debug(f"send_eval_{player_name}")
        
        for client_id in self.CONNECTION:
            player_name = self.CONNECTION[client_id]["name"]
            logger.debug(f"server_next_turn | waiting_recv_received_{player_name}")
            received = self.CONNECTION[client_id]["client_socket"].recv(16).decode(self.utf_8)
            logger.debug(f"server_next_turn | recv_received_{player_name}")
        
        if not skip:
            time.sleep(2)
            self.screen.widgets["Title"]["day"]["text"] = f"END OF DAY"
            self.screen.widgets["Title"]["demand"]["text"] = self.position
        
        if self.position < self.duration:
            self.screen.widgets["nextButton"].config(state=tk.NORMAL)
            self.screen.widgets["skipButton"].config(state=tk.NORMAL)
            self.server_wait_order() if not self.policy else None
        else:
            result_A = self.Inv.update_result("A")
            result_B = self.Inv.update_result("B")
            self.screen.game_over(
                A_game_var=-int(self.Inv.INV["A"]["cash_flow"]["Variable Cost"]),
                A_game_cycle=round(self.Inv.INV["A"]["num_order"]),
                A_true_var=-int(result_A["Total Variable Cost"]),
                A_true_cycle=round(result_A["Total number of orders"], 4),
                B_game_var=-int(self.Inv.INV["B"]["cash_flow"]["Variable Cost"]),
                B_game_cycle=round(self.Inv.INV["B"]["num_order"], 4),
                B_true_var=-int(result_B["Total Variable Cost"]),
                B_true_cycle=round(result_B["Total number of orders"].value, 4),
            ) if self.policy else None
        logger.debug(f"server_next_turn | self.position = {self.position}")
        
    def server_skip_turn(self, skip=0):
        self.screen.widgets["nextButton"].config(state=tk.DISABLED)
        self.screen.widgets["skipButton"].config(state=tk.DISABLED)
        self.screen.clear_order()
        
        
        for client_id in self.CONNECTION:
            player_name = self.CONNECTION[client_id]["name"]
            self.CONNECTION[client_id]["client_socket"].send("skip".encode(self.utf_8))
            logger.debug(f"server_skip_turn | send_command_skip_turn_{player_name}")
            
            self.receivethread[client_id].join()
            logger.debug(f"server_skip_turn | waiting_recv_ready_{player_name}")
            ready = self.CONNECTION[client_id]["client_socket"].recv(16).decode(self.utf_8)
            logger.debug(f"server_skip_turn | recv_ready_{player_name}")
            self.CONNECTION[client_id]["client_socket"].send(str(skip).encode(self.utf_8))
            logger.debug(f"server_skip_turn | send_skip_{player_name}")
        
        for client_id in self.CONNECTION:
            player_name = self.CONNECTION[client_id]["name"]
            logger.debug(f"server_skip_turn | waiting_recv_received server_xl : {player_name}")
        
        logger.debug(f"server_skip_turn | waiting_xl_{player_name}")
            # for i in range(self.duration-self.position):
            # 	Inv[player_name][(2+self.position+i, 11)].value = Inv[player_name][(2+self.position+i, 12)].value
        
        for client_id in self.CONNECTION:
            player_name = self.CONNECTION[client_id]["name"]
            logger.debug(f"server_skip_turn | waiting_recv_received_{player_name}")
            # Inv_data for get value or excel formula | Inv_data_only for get value without formula
    
            output_send = []
            for i in range(self.duration-self.position):
                if player_name == "A":
                    current_row_A = self.Inv.get_current_row("A")
                    data = {
                        'period': self.position,
                        'demand': current_row_A[self.Inv.col_dict_inv["demand"]],
                        'onhand': current_row_A[self.Inv.col_dict_inv["ending_inv"]],
                        'onorder': current_row_A[self.Inv.col_dict_inv["on_order"]],
                        'arrive': current_row_A[self.Inv.col_dict_inv["arrive"]],
                        'receive': current_row_A[self.Inv.col_dict_inv["recieve"]],
                        'sold': current_row_A[self.Inv.col_dict_inv["sold"]],
                        'Shortage': current_row_A[self.Inv.col_dict_inv["shortage"]],
                        'profits': self.Inv.INV[player_name]["cash_flow"]["Sales Profit"],
                        'varcost': -self.Inv.INV[player_name]["cash_flow"]["Variable Cost"],
                    }
                else:
                    current_row_B = self.Inv.get_current_row("A")
                    data = {
                        'period': self.position,
                        'demand': current_row_B[self.Inv.col_dict_inv["demand"]],
                        'onhand': current_row_B[self.Inv.col_dict_inv["ending_inv"]],
                        'onorder': current_row_B[self.Inv.col_dict_inv["on_order"]],
                        'arrive': current_row_B[self.Inv.col_dict_inv["arrive"]],
                        'receive': current_row_B[self.Inv.col_dict_inv["recieve"]],
                        'sold': current_row_B[self.Inv.col_dict_inv["sold"]],
                        'Shortage': current_row_B[self.Inv.col_dict_inv["shortage"]],
                        'profits': self.Inv.INV[player_name]["cash_flow"]["Sales Profit"],
                        'varcost': -self.Inv.INV[player_name]["cash_flow"]["Variable Cost"],
                    }
                output_send.append(data)
                

            output_send = str(output_send).encode('UTF-8')
            self.outputthread[client_id] = threading.Thread(
                target=self.CONNECTION[client_id]["client_socket"].send,
                args=(output_send,),
            )
        [self.outputthread[client_id].start() for client_id in self.outputthread]
        logger.debug(f"server_skip_turn | send_eval_list_{player_name}")
        
        for client_id in self.CONNECTION:
            player_name = self.CONNECTION[client_id]["name"]
            logger.debug(f"server_skip_turn | waiting_recv_received_{player_name}")
            received = self.CONNECTION[client_id]["client_socket"].recv(16).decode(self.utf_8)
            logger.debug(f"server_skip_turn | recv_received_{player_name}")
        
        for i in range(self.duration-self.position):
            self.screen.edit_score(
                self.position,
                "Ordering",
                A=[current_row_A[self.Inv.col_dict_inv["order"]]],
                B=[current_row_B[self.Inv.col_dict_inv["order"]]],
            ) if self.position - self.init_position else None
            
            self.screen.set_total(
                profit=[int(self.Inv.INV[team]["cash_flow"]["Sales Profit"]) for team in self.CONNECTION],
                # fix=[-int(self.Inv[team]["X12"].value) for team in self.CONNECTION],
                var=[-int(self.Inv.INV[team]["cash_flow"]["Variable Cost"]) for team in self.CONNECTION],
            )
            
            # self.Inv[self.sheet_config]["F15"].value = self.position+i
            column = [0,1,5,8,2,7,11,12,13,14,15]
            A = []
            B = []

            for col in column:
                A.append(current_row_A[col])
                B.append(current_row_B[col])
            self.screen.add_score(
                self.position+i+1,
                skip = skip,
                A = A,
                Main = [current_row_A[0],self.demand[-1]],
                B = B,
            )
            self.screen.update()
            time.sleep(.2)
        
        self.position = self.duration
        result_A = self.Inv.update_result("A")
        result_B = self.Inv.update_result("B")
        self.screen.game_over(
                A_game_var=-int(self.Inv.INV["A"]["cash_flow"]["Variable Cost"]),
                A_game_cycle=round(self.Inv.INV["A"]["num_order"]),
                A_true_var=-int(result_A["Total Variable Cost"]),
                A_true_cycle=round(result_A["Total number of orders"], 4),
                B_game_var=-int(self.Inv.INV["B"]["cash_flow"]["Variable Cost"]),
                B_game_cycle=round(self.Inv.INV["B"]["num_order"], 4),
                B_true_var=-int(result_B["Total Variable Cost"]),
                B_true_cycle=round(result_B["Total number of orders"].value, 4),
            ) if self.policy else None

    def server_skip(self):
        for _ in range(self.duration-self.position):
            self.server_next_turn(skip=1)

    def check_player_login(self):
        for client_id in self.CONNECTION:
            player_name = self.CONNECTION[client_id]["client_socket"].recv(16).decode(self.utf_8)
            self.check_players_name.append(player_name)

            logger.debug(f"server_set_game | send policy={self.screen.policy} to player_name={player_name} | self.CONNECTION[client_id]={self.CONNECTION[client_id]}")
            self.CONNECTION[client_id]["client_socket"].send(str(self.screen.policy).encode(self.utf_8))

        self.screen.mainscreen()
        self.server_wait_order()

    def server_set_game(self, val):
        
        self.position = self.init_position
        self.screen.policy = val
        
        F6 = self.Inv.config['holding_cost']
        F8 = self.Inv.config['lead_time']
        F2_F4 = [self.Inv.config['price'],self.Inv.config['item_cost'],self.Inv.config['order_cost'],self.Inv.config['holding_factor']]
        F11_F12 = [self.Inv.config['mu'],self.Inv.config['sigma']]
        logger.debug(f"server_set_game | F2_F4={F2_F4} | F6={F6} | F8={F8} | F11_F12={F11_F12}")

        param_val = [*F2_F4, F6, F8, tuple(F11_F12)]
        self.screen.set_parameters(
            **{
                param: param_val[i]
                for i, param in enumerate(self.screen.parameters)
            }
        )
        
        self.Inv.reset_db()


        # self.Inv[self.sheet_config]["C4"].value = f"={self.Inv['CONFIG']['F16'].value}"
        # self.Inv[self.sheet_config]["C4"].value = [[v] for v in self.Inv[self.sheet_config]["C4#"].value]
        
        logger.debug(f"server_set_game | self.screen.policy={self.screen.policy}")
        logger.debug(f"server_set_game | self.CONNECTION={self.CONNECTION}")

        for client_id in self.CONNECTION:
            player_name = self.CONNECTION[client_id]["name"]
            logger.debug(f"server_set_game | send policy={self.screen.policy} to player_name={player_name} | self.CONNECTION[client_id]={self.CONNECTION[client_id]}")
            self.CONNECTION[client_id]["client_socket"].send(str(self.screen.policy).encode(self.utf_8))
    
        for client_id in self.CONNECTION:
            player_name = self.CONNECTION[client_id]["name"]
            # Inv_data for get value or excel formula | Inv_data_only for get value without formula
            if self.screen.policy:
                self.Inv.INV[player_name]["policy"] = True
            else:
                self.Inv.INV[player_name]["policy"] = False

        
        self.screen.widgets["nextButton"].config(text="Start" if self.screen.policy else "Next", command=self.server_get_EOQ if self.screen.policy else self.server_next_turn)
        self.screen.widgets["skipButton"].config(command=self.server_skip_turn)
        
        Cost = self.screen.cost
        self.policy = self.screen.policy

        self.screen.mainscreen()
        self.server_wait_order()

    def server_get_EOQ(self):
        for client_id in self.CONNECTION:
            player_name = self.CONNECTION[client_id]["name"]
            # Inv_data for get value or excel formula | Inv_data_only for get value without formula
            # Inv_data, Inv_data_only = self.get_sheet_by_name(player_name)
            self.CONNECTION[client_id]["client_socket"].send("get_EOQ".encode(self.utf_8))
            logger.debug(f"send_command_get_EOQ_{player_name}")
            logger.debug(f"waiting_recv_EOQ_{player_name}")
            self.EOQ[player_name] = eval(self.CONNECTION[client_id]["client_socket"].recv(64).decode(self.utf_8))
            logger.debug(f"recv_EOQ_{player_name}")
            self.Inv.INV[player_name]['order_quantity']= int(self.EOQ[player_name]["EOQ"])
            self.Inv.INV[player_name]['reorder_point'] = int(self.EOQ[player_name]["ROP"])
            
            self.screen.widgets["nextButton"].config(text="Next", command=self.server_next_turn)
            self.screen.widgets["skipButton"].config(command=self.server_skip_turn)

    def server_restart(self):
        self.screen.widgets["nextButton"].config(state=tk.NORMAL)
        self.screen.widgets["skipButton"].config(state=tk.NORMAL)
        self.Inv.reset_db()
        self.Inv.reset()
        
        for client_id in self.CONNECTION:
            player_name = self.CONNECTION[client_id]["name"]
            # Inv_data for get value or excel formula | Inv_data_only for get value without formula
            self.CONNECTION[client_id]["client_socket"].send("restart".encode(self.utf_8))
            logger.debug(f"server_restart | send_command_restart_{player_name}")
            self.receivethread[client_id].join()
            logger.debug(f"server_restart | waiting_recv_ready_{player_name}")
            ready = self.CONNECTION[client_id]["client_socket"].recv(16).decode(self.utf_8)
            logger.debug(f"server_restart | recv_ready_{player_name}")
            self.screen.reset()

    def server_result(self, path):
        CONNECTION = ["A","B"]
        output_file(path)
        
        INV = xw.Book("save/data.xlsx").sheets
        data = {
            team:{
                "Day":[
                    INV[team][f"B3:B{self.duration+3}"].options(empty=float(0)).value[i//2]
                    for i in range(2*len(INV[team][f"B3:B{self.duration+3}"].options(empty=float(0)).value))
                    if not i%2 or INV[team][f"F4:F{self.duration+4}"].options(empty=float(0)).value[i//2] or INV[team][f"L2:L{self.duration+2}"].options(empty=float(0)).value[i//2]
                ],
                "Onhand":[
                    INV[team][f"J3:J{self.duration+3}"].options(empty=float(0)).value[i//2]
                    if not i%2 else INV[team][f"G4:G{self.duration+4}"].options(empty=float(0)).value[i//2]
                    for i in range(2*len(INV[team][f"B3:B{self.duration+3}"].options(empty=float(0)).value))
                    if not i%2 or INV[team][f"F4:F{self.duration+4}"].options(empty=float(0)).value[i//2] or INV[team][f"L2:L{self.duration+2}"].options(empty=float(0)).value[i//2]
                ],
                "OnOrder":[
                    INV[team][f"D3:D{self.duration+3}"].options(empty=float(0)).value[i//2]
                    if bool(i%2) == bool(INV[team][f"L2:L{self.duration+2}"].options(empty=float(0)).value[i//2]) else 0
                    for i in range(2*len(INV[team][f"B3:B{self.duration+3}"].options(empty=float(0)).value))
                    if not i%2 or INV[team][f"F4:F{self.duration+4}"].options(empty=float(0)).value[i//2] or INV[team][f"L2:L{self.duration+2}"].options(empty=float(0)).value[i//2]
                ],
            } for team in CONNECTION
        }

        line_source = {team:ColumnDataSource(data=data[team]) for team in CONNECTION}
        Inv_graph = figure(title="Inventory", height=500, width=1200)
        
        graph_lines = {
            team:Inv_graph.vline_stack(
                ["Onhand", "OnOrder"],
                x="Day",
                source = line_source[team],
                line_dash = ["solid", "dashed"],
                color = "Red" if team == "A" else "Blue"
            ) for team in CONNECTION
        }
        for team in CONNECTION:
            graph_lines[team][1].visible = False
        
        graph_labels = {
            f"{team}":Label(
                x = data[team]["Day"][-1],
                y = data[team]["Onhand"][-1],
                text = f"{team}_Inv"
            )
            for team in CONNECTION
        }

        graph_rop = {
            **{f"{team}_line":Inv_graph.line(x=[data[team]["Day"][0], data[team]["Day"][-1]], y=[self.EOQ[team]["ROP"], self.EOQ[team]["ROP"]], line_dash="dotted", color="Red" if team == self.sheet_a else "Blue") for team in self.CONNECTION},
            **{f"{team}_label":Label(x=data[team]["Day"][-1], y=self.EOQ[team]["ROP"], text=f"{team}_Reorder_Point") for team in CONNECTION},
        } if self.policy else None
        
        [Inv_graph.add_layout(graph_labels[team]) for team in graph_labels]
        [Inv_graph.add_layout(graph_rop[key]) for key in graph_rop] if graph_rop else None

        checkbox_team = CheckboxGroup(labels=["Team_A", "Team_B"], active=[0, 1], width=100)
        checkbox_lines = CheckboxGroup(labels=["On Hand", "Stock Position", "ROP"], active=[0, 2], width=100)
        check = CustomJS(
            args=dict(
                lines=graph_lines,
                labels=graph_labels,
                checkbox_team=checkbox_team,
                checkbox_lines=checkbox_lines,
                rop=graph_rop
            ),
            code="""
                for (let team of Object.keys(lines)) {
                    lines[team][0].visible = checkbox_team.active.includes(team == 'A' ? 0 : 1) && checkbox_lines.active.includes(0);
                    lines[team][1].visible = checkbox_team.active.includes(team == 'A' ? 0 : 1) && checkbox_lines.active.includes(1);
                }
                for (let team of Object.keys(labels)) {
                    labels[team].visible = checkbox_team.active.includes(team == 'A' ? 0 : 1) && (checkbox_lines.active.includes(0) || checkbox_lines.active.includes(1));
                }
                for (let key of Object.keys(rop)) {
                    rop[key].visible = checkbox_team.active.includes(key.split('_')[0] == 'A' ? 0 : 1) && checkbox_lines.active.includes(2);
                }
            """
        )
        
        checkbox_team.js_on_change("active", check)
        checkbox_lines.js_on_change("active", check)
        
        graph = row(Inv_graph, column(checkbox_team, checkbox_lines)) # (checkbox_head) Cost_graph, Profit_graph
        show(graph)
        
    def server_savefile(self):
        logger.debug(f"server_savefile | self.instance = {self.instance}")
        if not self.instance:
            logger.debug(f"server_savefile | Creating new directory...")
            tnow = time.strftime("%d-%h-%Y %H.%M.%S",time.localtime())
            logger.debug(f"server_savefile | {self.path_save}")
            path_directory = f"{self.path_save}/{tnow}"
            self.check_directory_is_delete(path_directory)

        self.instance = len(os.listdir(self.path_save)) if self.instance is None else self.instance
        logger.debug(f"server_savefile | self.instance = {self.instance}")

        # path 
        # directory = len(os.listdir(f"save/{tnow}"))
        # file_policy = "EOQ_Policy" if self.policy else "No_Policy"
        # path_file_policy= f"{self.path_save}/{tnow}/{directory}_{file_policy}"
        # self.check_directory_is_create(path_file_policy)

        new_workbook = openpyxl.Workbook()
        excel_keys = ["B","C","D","E","F","G","H","I","J","L","M","O","P","Q","R","S","T","U"]
        for i,team in enumerate(["A","B"]):
            team_values = self.Inv.get_all_data(team)
            new_worksheet = new_workbook.active if i == 0 else new_workbook.create_sheet()
            new_worksheet.title = f"{team}"
            for key in self.Inv.excel_column.keys():
                new_worksheet[self.Inv.excel_column[key] + "2"].value = key
            for row in range(len(team_values)):
                for col in range(len(excel_keys)):
                    new_worksheet[excel_keys[col] + str(row+4)].value = team_values[row][col]
            new_worksheet["W10"].value = "DETAIL OF CASH FLOW"
            #row for write result in player sheet
            count = 2
            keys_player = ["fill_rate","service_level","order_quantity","reorder_point","policy","cash_flow"]
            for i in range(len(keys_player)):
                if keys_player[i] != "cash_flow":
                    new_worksheet[f"W{count}"] = keys_player[i]
                    new_worksheet[f"X{count}"] = self.Inv.INV[team][keys_player[i]]
                    count+=1
                    if count in [4,7,9] :
                        count+=1
                else:
                    count+=1
                    for key in self.Inv.INV[team][keys_player[i]]:
                        new_worksheet[f"W{count}"] = key
                        new_worksheet[f"X{count}"] = self.Inv.INV[team][keys_player[i]][key]
                        count+=1
            new_worksheet["Z11"] = "Total number of units sold"
            new_worksheet["Z12"] = "Total number of units purchased"
            new_worksheet["Z13"] = "Total number of orders"
            new_worksheet["Z14"] = "Total number of shortages"
            new_worksheet["Z15"] = "Total number of leftovers"

            new_worksheet["AA11"] = self.Inv.INV[team]["unit_sold"]
            new_worksheet["AA12"] = self.Inv.INV[team]["unit_purchase"]
            new_worksheet["AA13"] = self.Inv.INV[team]["num_order"]
            new_worksheet["AA14"] = self.Inv.INV[team]["shortages"]
            new_worksheet["AA15"] = self.Inv.INV[team]["leftover"]


        new_workbook.save(os.path.abspath(f"save/data.xlsx"))
        self.server_result(os.path.abspath(f"save/graph.html"))

    def clear_contents(self, sheet, columns):
        # # Assuming self.Inv[sheet][columns] returns a range of cells
        # for cell in self.Inv[sheet][columns]:
        #     cell.value = None
        pass
    def check_directory_is_create(self, path_directory):
        if not os.path.exists(path_directory):
            os.makedirs(path_directory)
            logger.debug("check_directory_is_create | Directory created:", path_directory)
        else:
            logger.debug("check_directory_is_create | Directory already exists:", path_directory)

    def check_directory_is_delete(self, path_directory):
        if os.path.exists(path_directory):
            os.mkdir(os.path.abspath(f"{path_directory}"))
            logger.debug("check_directory_is_delete | Directory deleted:", path_directory)
        else:
            logger.debug("check_directory_is_delete | Directory not found:", path_directory)