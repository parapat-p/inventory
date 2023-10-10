import os
import openpyxl
from openpyxl import Workbook
from loguru import logger

class ExcelDatabase(object):
    def __init__(self):
        self.directory = os.path.join(".", "save")
        self.filename = "data.xlsx"
        self.path_db = os.path.join(self.directory, self.filename)
        
    def open_excel(self):
        inventory = None
        __exception = None
        try:
            if not os.path.exists(self.directory):
                # Create the directory if it doesn't exist
                os.makedirs(self.directory)

            if not os.path.exists(self.path_db):
                # Create a new workbook if the file doesn't exist
                self.workbook = Workbook()
                self.workbook.save(self.path_db)
                self.workbook.close()

            self.workbook = openpyxl.load_workbook(self.path_db, data_only=False)
            self.workbook_data_only = openpyxl.load_workbook(self.path_db, data_only=True)

            sheet_config = self.workbook.get_sheet_by_name("CONFIG")
            sheet_a = self.workbook.get_sheet_by_name("A")
            sheet_b = self.workbook.get_sheet_by_name("B")
            sheet_result = self.workbook.get_sheet_by_name("RESULT")

            sheet_config_data_only = self.workbook_data_only.get_sheet_by_name("CONFIG")
            sheet_a_data_only = self.workbook_data_only.get_sheet_by_name("A")
            sheet_b_data_only = self.workbook_data_only.get_sheet_by_name("B")
            sheet_result_data_only = self.workbook_data_only.get_sheet_by_name("RESULT")

        
        except Exception as e:
            logger.error(e)
            __exception = e

        finally:
            if __exception:
                return __exception
            if sheet_a and sheet_b:
                inventory = {
                    "A": sheet_a,
                    "B": sheet_b,
                }
                inventory_data_only = {
                    "A": sheet_a_data_only,
                    "B": sheet_b_data_only,
                }
                return inventory, inventory_data_only
